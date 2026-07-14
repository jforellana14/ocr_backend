from __future__ import annotations

import os
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import Engine


# ---------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent

XLSX_PATH = (
    BASE_DIR
    / "imports"
    / "precios_mem_2026.xlsx"
)

START_DATE = date(2026, 5, 1)
END_DATE = date.today() - timedelta(days=1)

SOURCE_NAME = "MEM"
NOTES = "Carga histórica única desde Excel oficial del MEM."


# ---------------------------------------------------------------------
# NORMALIZACIÓN
# ---------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    result = str(value or "").strip().upper()

    result = unicodedata.normalize("NFKD", result)
    result = "".join(
        character
        for character in result
        if not unicodedata.combining(character)
    )

    result = re.sub(r"\s+", " ", result)

    return result.strip()


def parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    raw = str(value).strip()

    if not raw:
        return None

    # ISO con hora.
    if "T" in raw:
        raw = raw.split("T", maxsplit=1)[0]

    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d/%m/%y",
        "%d-%m-%y",
    )

    for date_format in formats:
        try:
            return datetime.strptime(
                raw[:10],
                date_format,
            ).date()
        except ValueError:
            continue

    return None


def parse_price(value: Any) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        price = float(value)
    else:
        raw = (
            str(value)
            .replace("Q.", "")
            .replace("Q", "")
            .replace(",", "")
            .strip()
        )

        try:
            price = float(raw)
        except (TypeError, ValueError):
            return None

    # Validación de seguridad.
    if not 5.00 <= price <= 100.00:
        return None

    return round(price, 2)


# ---------------------------------------------------------------------
# LECTURA DEL EXCEL
# ---------------------------------------------------------------------

def find_columns(
    worksheet,
) -> tuple[int, int, int]:
    """
    Detecta automáticamente:

    - fila de encabezados;
    - columna FECHA;
    - columna DIÉSEL.
    """

    maximum_header_rows = min(
        worksheet.max_row,
        80,
    )

    for row_number in range(
        1,
        maximum_header_rows + 1,
    ):
        date_column = None
        diesel_column = None

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            value = normalize_text(
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
            )

            if (
                value == "FECHA"
                or value.startswith("FECHA ")
                or "FECHA DE MONITOREO" in value
            ):
                date_column = column_number

            if (
                "DIESEL" in value
                or "ACEITE COMBUSTIBLE DIESEL" in value
                or "COMBUSTIBLE DIESEL" in value
            ):
                diesel_column = column_number

        if date_column and diesel_column:
            return (
                row_number,
                date_column,
                diesel_column,
            )

    raise RuntimeError(
        f"No se encontraron las columnas FECHA y DIÉSEL "
        f"en la hoja {worksheet.title!r}."
    )


def extract_records() -> list[tuple[date, float]]:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            "No se encontró el Excel del MEM:\n"
            f"{XLSX_PATH}"
        )

    workbook = load_workbook(
        XLSX_PATH,
        data_only=True,
        read_only=True,
    )

    records: dict[date, float] = {}
    processed_sheets = 0

    try:
        for worksheet in workbook.worksheets:
            try:
                (
                    header_row,
                    date_column,
                    diesel_column,
                ) = find_columns(worksheet)

            except RuntimeError:
                print(
                    f"Hoja ignorada: {worksheet.title!r}"
                )
                continue

            processed_sheets += 1

            print(
                f"Procesando hoja {worksheet.title!r}: "
                f"encabezado={header_row}, "
                f"fecha={date_column}, "
                f"diésel={diesel_column}"
            )

            for row_number in range(
                header_row + 1,
                worksheet.max_row + 1,
            ):
                record_date = parse_date(
                    worksheet.cell(
                        row=row_number,
                        column=date_column,
                    ).value
                )

                diesel_price = parse_price(
                    worksheet.cell(
                        row=row_number,
                        column=diesel_column,
                    ).value
                )

                if (
                    record_date is None
                    or diesel_price is None
                ):
                    continue

                if not (
                    START_DATE
                    <= record_date
                    <= END_DATE
                ):
                    continue

                records[record_date] = diesel_price

    finally:
        workbook.close()

    if processed_sheets == 0:
        raise RuntimeError(
            "Ninguna hoja del Excel contiene las columnas "
            "FECHA y DIÉSEL."
        )

    if not records:
        raise RuntimeError(
            "No se encontraron precios válidos entre "
            f"{START_DATE} y {END_DATE}."
        )

    return sorted(
        records.items(),
        key=lambda item: item[0],
    )


# ---------------------------------------------------------------------
# VALIDACIÓN DE COBERTURA
# ---------------------------------------------------------------------

def expected_dates() -> set[date]:
    values: set[date] = set()
    current = START_DATE

    while current <= END_DATE:
        values.add(current)
        current += timedelta(days=1)

    return values


def show_preview(
    records: list[tuple[date, float]],
) -> None:
    found_dates = {
        record_date
        for record_date, _ in records
    }

    missing_dates = sorted(
        expected_dates() - found_dates
    )

    print("\n" + "=" * 64)
    print("VISTA PREVIA DEL HISTÓRICO")
    print("=" * 64)

    print(
        f"Rango solicitado: "
        f"{START_DATE} al {END_DATE}"
    )
    print(
        f"Primer registro encontrado: "
        f"{records[0][0]} — Q{records[0][1]:.2f}"
    )
    print(
        f"Último registro encontrado: "
        f"{records[-1][0]} — Q{records[-1][1]:.2f}"
    )
    print(
        f"Registros encontrados: {len(records)}"
    )
    print(
        f"Fechas sin registro: {len(missing_dates)}"
    )

    print("\nPrimeros cinco registros:")

    for record_date, price in records[:5]:
        print(
            f"  {record_date.isoformat()} — Q{price:.2f}"
        )

    print("\nÚltimos cinco registros:")

    for record_date, price in records[-5:]:
        print(
            f"  {record_date.isoformat()} — Q{price:.2f}"
        )

    if missing_dates:
        print("\nFechas faltantes:")

        for missing_date in missing_dates[:20]:
            print(f"  {missing_date.isoformat()}")

        if len(missing_dates) > 20:
            print(
                f"  ... y {len(missing_dates) - 20} más."
            )

        print(
            "\nEl script no inventará valores para esas fechas."
        )


# ---------------------------------------------------------------------
# POSTGRESQL
# ---------------------------------------------------------------------

def get_database_url() -> str:
    database_url = os.getenv(
        "DATABASE_URL",
        "",
    ).strip()

    if not database_url:
        raise RuntimeError(
            "Falta la variable de entorno DATABASE_URL."
        )

    if database_url.startswith("postgres://"):
        database_url = database_url.replace(
            "postgres://",
            "postgresql://",
            1,
        )

    return database_url


def create_database_engine() -> Engine:
    return create_engine(
        get_database_url(),
        pool_pre_ping=True,
    )


def get_fuel_price_columns(
    engine: Engine,
) -> set[str]:
    inspector = inspect(engine)

    if not inspector.has_table("fuel_prices"):
        raise RuntimeError(
            "No existe la tabla fuel_prices."
        )

    return {
        column["name"]
        for column in inspector.get_columns(
            "fuel_prices"
        )
    }


def validate_database_schema(
    columns: set[str],
) -> None:
    required_columns = {
        "fecha",
        "precio_galon",
    }

    missing = required_columns - columns

    if missing:
        raise RuntimeError(
            "La tabla fuel_prices no tiene las columnas "
            "obligatorias: "
            + ", ".join(sorted(missing))
        )


def build_upsert_query(
    available_columns: set[str],
):
    """
    Adapta el INSERT a la estructura real de fuel_prices.

    Las columnas fuente, observaciones y created_at son opcionales.
    """

    insert_columns = [
        "fecha",
        "precio_galon",
    ]

    value_fields = [
        ":fecha",
        ":precio_galon",
    ]

    update_fields = [
        "precio_galon = EXCLUDED.precio_galon",
    ]

    if "fuente" in available_columns:
        insert_columns.append("fuente")
        value_fields.append(":fuente")
        update_fields.append(
            "fuente = EXCLUDED.fuente"
        )

    if "observaciones" in available_columns:
        insert_columns.append("observaciones")
        value_fields.append(":observaciones")
        update_fields.append(
            "observaciones = EXCLUDED.observaciones"
        )

    if "created_at" in available_columns:
        insert_columns.append("created_at")
        value_fields.append("NOW()")

    query = f"""
        INSERT INTO fuel_prices (
            {", ".join(insert_columns)}
        )
        VALUES (
            {", ".join(value_fields)}
        )
        ON CONFLICT (fecha)
        DO UPDATE SET
            {", ".join(update_fields)}
    """

    return text(query)


def save_records(
    engine: Engine,
    records: list[tuple[date, float]],
) -> dict[str, int]:
    columns = get_fuel_price_columns(engine)
    validate_database_schema(columns)

    existing_query = text(
        """
        SELECT precio_galon
        FROM fuel_prices
        WHERE fecha = :fecha
        """
    )

    upsert_query = build_upsert_query(
        columns
    )

    inserted = 0
    updated = 0
    unchanged = 0

    with engine.begin() as connection:
        for record_date, price in records:
            previous_price = connection.execute(
                existing_query,
                {
                    "fecha": record_date,
                },
            ).scalar_one_or_none()

            if previous_price is None:
                inserted += 1

            elif (
                round(float(previous_price), 2)
                == price
            ):
                unchanged += 1

            else:
                updated += 1

            parameters = {
                "fecha": record_date,
                "precio_galon": price,
                "fuente": SOURCE_NAME,
                "observaciones": NOTES,
            }

            connection.execute(
                upsert_query,
                parameters,
            )

    return {
        "insertados": inserted,
        "actualizados": updated,
        "sin_cambios": unchanged,
    }


# ---------------------------------------------------------------------
# EJECUCIÓN
# ---------------------------------------------------------------------

def main() -> int:
    print("=" * 64)
    print("IMPORTACIÓN HISTÓRICA DE DIÉSEL — MEM → POSTGRESQL")
    print("=" * 64)

    try:
        records = extract_records()

        show_preview(records)

        confirmation = input(
            "\nEscribe SI para cargar estos registros "
            "en PostgreSQL: "
        ).strip().upper()

        if confirmation != "SI":
            print(
                "\nOperación cancelada. "
                "No se modificó la base de datos."
            )
            return 0

        engine = create_database_engine()

        result = save_records(
            engine=engine,
            records=records,
        )

        print("\n" + "=" * 64)
        print("IMPORTACIÓN COMPLETADA")
        print("=" * 64)

        print(
            f"Registros insertados:   "
            f"{result['insertados']}"
        )
        print(
            f"Registros actualizados: "
            f"{result['actualizados']}"
        )
        print(
            f"Registros sin cambios:  "
            f"{result['sin_cambios']}"
        )
        print(
            f"Total procesado:         "
            f"{len(records)}"
        )

        return 0

    except Exception as error:
        print("\n" + "=" * 64)
        print("ERROR")
        print("=" * 64)
        print(error)

        return 1


if __name__ == "__main__":
    sys.exit(main())