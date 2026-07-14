import os
import json
import tempfile
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from database import get_db
from models import User, TariffImportHistory
from auth import require_roles
from app.services.tariff_import_service import (
    import_tariff_rows_and_update_documents,
)

router = APIRouter(
    prefix="/admin/tariff-import",
    tags=["Admin Tariff Import"]
)


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def money(value):
    if value is None or value == "":
        return None

    text = str(value).replace("Q", "").replace(",", "").strip()

    try:
        return float(text)
    except Exception:
        return None


def read_excel_rows(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        origen = clean(row[0])
        destino = clean(row[1])

        if not origen or not destino:
            continue

        prices = [
            money(row[4]),
            money(row[5]),
            money(row[6]),
            money(row[7]),
            money(row[8]),
            money(row[9]),
            money(row[10]),
            money(row[11]),
        ]

        if any(price is None for price in prices):
            continue

        rows.append({
            "origen": origen,
            "destino": destino,
            "kilometraje": money(row[2]),
            "consumo_diesel": money(row[3]),
            "prices": prices
        })

    return rows


@router.post("/excel")
async def import_tariff_excel(
    file: UploadFile = File(...),
    force_recalculate: bool = False,
    dry_run: bool = True,
    version: str = "2026",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN"))
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Debe subir un archivo Excel .xlsx"
        )

    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)

    with open(file_path, "wb") as buffer:
        buffer.write(await file.read())

    rows = read_excel_rows(file_path)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron filas válidas en el Excel."
        )

    if dry_run:
        return {
            "modo": "SIMULACION",
            "archivo": file.filename,
            "filas_validas": len(rows),
            "mensaje": "Simulación completada. No se guardó ningún dato."
        }

    result = import_tariff_rows_and_update_documents(
        db=db,
        rows=rows,
        force_recalculate=force_recalculate
    )

    history = TariffImportHistory(
        archivo=file.filename,
        usuario=current_user.username,
        version=version,
        filas=len(rows),
        rutas_creadas=result.get("created_routes", 0),
        tarifarios_creados=result.get("created_plans", 0),
        rangos_creados=result.get("created_details", 0),
        viajes_actualizados=result.get("updated_documents", 0),
        errores=json.dumps(result.get("failed_documents", []), ensure_ascii=False),
        estado="COMPLETADO"
    )

    db.add(history)
    db.commit()
    db.refresh(history)

    result["filas_leidas"] = len(rows)
    result["history_id"] = history.id

    return result


@router.get("/history")
def get_tariff_import_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("ADMIN", "ENCARGADO"))
):
    return (
        db.query(TariffImportHistory)
        .order_by(TariffImportHistory.fecha_importacion.desc())
        .all()
    )